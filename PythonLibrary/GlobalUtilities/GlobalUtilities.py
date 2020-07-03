from openpyxl import *


class GlobalUtilities():

    def __init__(self):
        pass

    def getExcelInDictionary(self,workBookpath, sheetName, TestCaseName):
        """
        Gets Data from Excel to Dictionary
        :param workBookpath:
        :param sheetName:
        :param TestCaseName:
        :return:
        """
        wb = load_workbook(workBookpath, read_only=True)
        ws = wb.get_sheet_by_name(sheetName)

        cols = ws.max_column
        rows = ws.max_row

        # Find Test Case Sequential Search.. If Data Grows more than 1000 cols shift to Binary Search
        i = 0
        for i in range(1, cols + 1):
            if ws.cell(row=1, column=i).value == TestCaseName:
                break

        if 1 <= i <= cols:
            data = {}
            for rc in range(1, rows + 1):
                if ws.cell(row=rc, column=1).value is not None:
                    data[ws.cell(row=rc, column=1).value] = ws.cell(row=rc, column=i).value
                else:
                    break
            wb.close()
            return data
        else:
            print("Data Not Found.Please Check the Test Case Name is Valid and Re-Try")
            wb.close()
            return -1